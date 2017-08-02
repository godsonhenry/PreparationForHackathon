# file name is readwrite
from openpyxl import load_workbook, Workbook
from season import TeamsList, GamesList 
import datetime
def read_teams(filename):
    wb = load_workbook(filename = filename)
    ws = wb['Division_Info']
    teamlist = TeamsList()
    for i in range(2,32):
        teamlist.add(ws[i][0].value, ws[i][1].value, ws[i][2].value)
    return teamlist

def read_games(filename):
    wb = load_workbook(filename = filename)
    ws = wb['2016_17_NBA_Scores']
    gamelist = GamesList()
    for i in range(2,1232):
        gamelist.add(ws[i][0].value, ws[i][1].value, ws[i][2].value,
                     ws[i][5].value)
    return gamelist

def write_outlist(outlist):
    wb = Workbook()
    ws = wb.create_sheet('NBA_Clinch_Dates')
    i = 0
    ws.cell(row=1, column=1, value='Team')
    ws.cell(row=1, column=2, value='Date Eliminated')
    for ele in outlist:
        i += 1
        ws.cell(row=i+2, column=1, value=ele)
        ws.cell(row=i+2, column=2, value=outlist[ele])
    wb.save('result.xlsx')

if __name__ == '__main__':
    teamlist = read_teams('Analytics_Attachment.xlsx')
    #teamlist.print()
    gamelist = read_games('Analytics_Attachment.xlsx')
    #gamelist.print() 

