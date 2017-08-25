# The answer for the question is in code using Separator to highlight.
# just like this 
#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
'''
what I did here is that I calculate all the possible probability in each 
round for each team. So if I have all the probability, I could do any
questions, just limited by time.
By the way, all possible results is stored in a 4*16*16*8 matrixs
as for 4 round * 16 teams * 16 teams (against with each other) * 8 
outcomes (wins in 4,5,6,7 games, and the winner could be either)
'''


import numpy as np
from random import random
from openpyxl import load_workbook, Workbook
import matplotlib.pyplot as plt


wb = load_workbook(filename = 'win_probabilities.xlsx')
ws = wb['win_probabilities.csv']
wins = np.ndarray(shape=(16,16), dtype=float)
for i in range(16):
    for j in range(16):
        wins[i][j] = 0.0
i = 1

# west team will have small numbers

while i < len(ws['A']):
    if ws['A'][i].value[0] == 'W':
        home = int(ws['A'][i].value[4]) - 1
    else:
        home = int(ws['A'][i].value[4]) + 7
    if ws['B'][i].value[0] == 'W':
        away = int(ws['B'][i].value[4]) - 1
    else:
        away = int(ws['B'][i].value[4]) + 7
    wins[home][away] = ws['C'][i].value
    wins[away][home] = 1-ws['D'][i].value
    i += 1




wb = load_workbook(filename = 'Business-Track-Application-Datasets.xlsx')
ws = wb['Hypothetical Playoff Gate Data']
rev = np.ndarray(shape=(16,4), dtype=float)
for i in range(3, 19):
    for j in range(2, 6):
        if i >= 11:
            k = i - 8
        else:
            k = i + 8
        rev[k-3][j-2] = ws[i][j].value


def t2wins8(ph, pa):
    pa = 1 - pa
    w4 = ph**2 * pa**2
    w5 = (2*ph*(1-ph)*pa**2+2*ph**2*pa*(1-pa))*ph
    w6 = (3*ph*(1-ph)**2*pa**2+ph**3*(1-pa)**2+3*2*ph**2*(1-ph)*pa*(1-pa))*pa
    w7 = ((1-ph)**3*pa**3
          + ph**3*(1-pa)**3
          + 9*ph**2*(1-ph)*pa*(1-pa)**2
          + 9*ph*(1-ph)**2*pa**2*(1-pa)) * ph
    ph, pa = 1-ph, 1-pa
    l4 = ph**2 * pa**2
    l5 = (2*ph*(1-ph)*pa**2+2*ph**2*pa*(1-pa))*ph
    l6 = (3*ph*(1-ph)**2*pa**2+ph**3*(1-pa)**2+3*2*ph**2*(1-ph)*pa*(1-pa))*pa
    l7 = ((1-ph)**3*pa**3
          + ph**3*(1-pa)**3
          + 9*ph**2*(1-ph)*pa*(1-pa)**2
          + 9*ph*(1-ph)**2*pa**2*(1-pa)) * ph    
    
    return [w4, w5, w6, w7, l4, l5, l6, l7]

# get a op[i][j]
# it stores the teams that in i round the team j probably will against with

from copy import deepcopy
i = 0
op = []
a0 = []
for j in range(8):
    a0.append([7-j])
for j in range(8):
    a0.append([15-j])
op.append(a0)
a1 = [[3,4],[2,5],[1,6],[0,7],[0,7],[1,6],[2,5],[3,4]]
b1 = deepcopy(a1)
for i in range(8):
    for j in range(2):
        b1[i][j] = a1[i][j] + 8
for ele in b1:
    a1.append(ele)
op.append(a1)
a2 = [[1,2,5,6],[0,3,4,7],[0,3,4,7],[1,2,5,6],[1,2,5,6],[0,3,4,7],[0,3,4,7],[1,2,5,6]]
b2 = deepcopy(a2)
for i in range(8):
    for j in range(4):
        b2[i][j] = a2[i][j] + 8
for ele in b2:
    a2.append(ele)
op.append(a2)
a3 = [[8,9,10,11,12,13,14,15],[8,9,10,11,12,13,14,15],[8,9,10,11,12,13,14,15],[8,9,10,11,12,13,14,15],[8,9,10,11,12,13,14,15],[8,9,10,11,12,13,14,15],[8,9,10,11,12,13,14,15],[8,9,10,11,12,13,14,15]]
b3 = deepcopy(a3)
for i in range(8):
    for j in range(8):
        b3[i][j] = a3[i][j] - 8
for ele in b3:
    a3.append(ele)
op.append(a3)


# init

round_results = np.ndarray(shape=(4,16,16,8), dtype=float)
enter_pro = np.ndarray(shape=(4,16), dtype=float)
for i in range(4):
    for ii in range(16):
        for iii in range(16):
            for iiii in range(8):
                round_results[i,ii,iii,iiii] = 0.0
for i in range(4):
    for ii in range(16):
        enter_pro[i,ii] = 0.0
for ii in range(16):
    enter_pro[0,ii] = 1.0


# calculate all the possible results in each round for each team

for i in range(4):
    for ii in range(16):
        for iii in op[i][ii]:
            if not((iii < ii and ii < 8) or (iii >= 8 and ii > iii) or (iii + 8 < ii)):
                round_results[i,ii,iii] = (t2wins8(wins[ii][iii],wins[iii][ii]))
                for iiii in range(8):
                    round_results[i,ii,iii,iiii] = round_results[i,ii,iii,iiii] * enter_pro[i][ii] * enter_pro[i][iii]
    if i == 3:
        break
    for ii in range(16):
        for iii in range(16):
            for iiii in range(4):
                if not((iii < ii and ii < 8) or (iii >= 8 and ii > iii) or (iii + 8 < ii)):
                    enter_pro[i+1][ii] += round_results[i,ii,iii,iiii]
                    
                else:
                    enter_pro[i+1][ii] += round_results[i,iii,ii,iiii+4]

    
    

#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------

# 0 8 meet in final probablity is :


#  p = enter_pro[3][0] * enter_pro[3][8] 




# 0 8 meet in finals 

round_results = np.ndarray(shape=(4,16,16,8), dtype=float)
enter_pro = np.ndarray(shape=(4,16), dtype=float)
for i in range(4):
    for ii in range(16):
        for iii in range(16):
            for iiii in range(8):
                round_results[i,ii,iii,iiii] = 0.0
for i in range(4):
    for ii in range(16):
        enter_pro[i,ii] = 0.0
for ii in range(16):
    enter_pro[0,ii] = 1.0


for i in range(3):
    for ii in range(16):
        for iii in op[i][ii]:
            if not((iii < ii and ii < 8) or (iii >= 8 and ii > iii) or (iii + 8 < ii)):
                round_results[i,ii,iii] = (t2wins8(wins[ii][iii],wins[iii][ii]))
                if ii==0 or ii==8:
                    sums = 0
                    for iiii in range(4):
                        sums += round_results[i,ii,iii,iiii]
                    for iiii in range(4):
                        round_results[i,ii,iii,iiii] = round_results[i,ii,iii,iiii] / sums
                    for iiii in range(4):
                        round_results[i,ii,iii,iiii+4] = 0.0
            for iiii in range(8):
                round_results[i,ii,iii,iiii] = round_results[i,ii,iii,iiii] * enter_pro[i][ii] * enter_pro[i][iii]
    for ii in range(16):
        for iii in range(16):
            for iiii in range(4):
                if not((iii < ii and ii < 8) or (iii >= 8 and ii > iii) or (iii + 8 < ii)):
                    enter_pro[i+1][ii] += round_results[i,ii,iii,iiii]
                else:
                    enter_pro[i+1][ii] += round_results[i,iii,ii,iiii+4]
                    
for i in range(3,4):
    for ii in range(16):
        for iii in op[i][ii]:
            if not((iii < ii and ii < 8) or (iii >= 8 and ii > iii) or (iii + 8 < ii)):
                round_results[i,ii,iii] = (t2wins8(wins[ii][iii],wins[iii][ii]))
            for iiii in range(8):
                round_results[i,ii,iii,iiii] = round_results[i,ii,iii,iiii] * enter_pro[i][ii] * enter_pro[i][iii]
                
                
#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
# this is the total revenue when East1 meets West1 in the finals

total = 0
for i in range(4):
    for ii in range(16):
        for iii in range(16):
            if iii > ii:
                w4 = round_results[i,ii,iii,0] + round_results[i,ii,iii,4]
                w5 = round_results[i,ii,iii,1] + round_results[i,ii,iii,5]
                w6 = round_results[i,ii,iii,2] + round_results[i,ii,iii,6]
                w7 = round_results[i,ii,iii,7] + round_results[i,ii,iii,7]
                r4 = w4 * (2*rev[ii][i]+2*rev[iii][i])
                r5 = w5 * (3*rev[ii][i]+2*rev[iii][i])
                r6 = w6 * (3*rev[ii][i]+3*rev[iii][i])
                r7 = w7 * (4*rev[ii][i]+4*rev[iii][i])
                total = total + r4 + r5 + r6 + r7
#
#total

#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
#----------------------------------------------------------------
'''
Due to time limits, I could not write the answer for the last two 
questions, it is too complicated. However, I could give my idea.

From what we did above, we get all the possible results from the 
teams. And we can make change who wins or lose, when it is necessary.
So if we want to know W1 doesn’t meet E1, I could calculate the 
probability that W1 lose in the first round, but W1 wins in the 
first round. And let W1 loses and E1 wins so that I could get a 
expected revenue in this situation. So I could have one probability
and the expected revenue of that probability.
So we now need 8 more. They are:
W1 wins in 1st round, not E1.
W1 lose in 1st round, and E1.
Both wins in 1st round, W1 lose in 2nd, not E1
Both wins in 1st round, E1 lose in 2nd, not w1
Both wins in 1st round, both lose in 2nd
Both wins in 1st round, both wins in 2st round, W1 lose in 3rd, not E1
Both wins in 1st round, both wins in 2st round, E1 lose in 3rd, not W1
Both wins in 1st round, both wins in 2st round, both lose in 3rd
I could have all the probability and expected revenue of them. So I
could calculate the results.

As for how E5 defeat E4 in the 1st round will change the result. I just
simply need to change that game to a certain case, the rest will be the
same.

'''
